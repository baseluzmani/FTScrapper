# HSBCpensionimport.py
# Imports fund data from the "HSBC Pension" sheet in Funds_Database.xlsx
# Can be run repeatedly — only new rows (after the latest saved date) are inserted.
#
# Expected columns in the sheet:
#   Fund Code | Fund Name | Unit Price Date | Currency Code | Unit Price
#
# Rules:
#   - Fund ID   = "FUNDCODE:CURRENCY"  (e.g. HAPE:GBP)
#   - Fund Name = "FUNDCODE - Fund Name"
#   - No open/high/low — all set to same value as close (Unit Price)
#   - Volume = 0
#
# Usage:
#   python3 HSBCpensionimport.py                         # reads Funds_Database.xlsx
#   python3 HSBCpensionimport.py ~/Downloads/myfile.xlsx # specific file

import sys
import openpyxl
from datetime import datetime
import database

DEFAULT_FILE = "Funds_Database.xlsx"
DEFAULT_SHEET = "HSBC Pension"


def parse_excel(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"  ERROR: Sheet '{sheet_name}' not found.")
        print(f"  Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]

    # Read header row to find column positions dynamically
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    print(f"  Sheet: '{sheet_name}'")
    print(f"  Columns: {headers}")

    # Find column indices by name
    def find_col(name):
        for i, h in enumerate(headers):
            if name.lower() in h.strip().lower():
                return i
        return None

    col_code = find_col("fund code")
    col_name = find_col("fund name")
    col_date = find_col("unit price date")
    col_currency = find_col("currency code")
    col_price = find_col("unit price")

    # unit price date also matches unit price — make sure price col is exact match
    if col_price == col_date:
        for i, h in enumerate(headers):
            if h.strip().lower() == "unit price":
                col_price = i
                break

    missing = [
        n
        for n, c in [
            ("Fund Code", col_code),
            ("Fund Name", col_name),
            ("Unit Price Date", col_date),
            ("Currency Code", col_currency),
            ("Unit Price", col_price),
        ]
        if c is None
    ]

    if missing:
        print(f"  ERROR: Could not find columns: {missing}")
        sys.exit(1)

    # Group rows by fund
    funds = {}
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fund_code = str(row[col_code]).strip() if row[col_code] else None
        fund_name = str(row[col_name]).strip() if row[col_name] else None
        date_val = row[col_date]
        currency = str(row[col_currency]).strip() if row[col_currency] else "GBP"
        price = row[col_price]

        # Skip empty rows
        if not fund_code or not date_val or price is None:
            skipped += 1
            continue

        # Build fund ID and display name
        fund_id = f"{fund_code}:{currency}"
        full_name = f"{fund_code} - {fund_name}"

        # Parse date — handles datetime objects and strings
        if isinstance(date_val, datetime):
            date_formatted = date_val.strftime("%Y-%m-%d")
        else:
            try:
                date_formatted = datetime.strptime(
                    str(date_val).strip(), "%d/%m/%Y"
                ).strftime("%Y-%m-%d")
            except ValueError:
                try:
                    date_formatted = datetime.strptime(
                        str(date_val).strip(), "%Y-%m-%d"
                    ).strftime("%Y-%m-%d")
                except ValueError:
                    print(f"  WARNING: Could not parse date '{date_val}', skipping.")
                    skipped += 1
                    continue

        # Skip rows where price is not a number
        try:
            price = float(price)
        except (TypeError, ValueError):
            skipped += 1
            continue

        if fund_id not in funds:
            funds[fund_id] = {"name": full_name, "rows": []}

        funds[fund_id]["rows"].append(
            {
                "date": date_formatted,
                "open": price,
                "high": price,
                "low": price,
                "close": price,
                "volume": 0,
            }
        )

    return funds, skipped


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    print(f"Reading: {filepath}")

    funds, skipped = parse_excel(filepath, DEFAULT_SHEET)

    if not funds:
        print("  No valid data found. Exiting.")
        sys.exit(1)

    conn = database.get_connection()
    database.create_table(conn)

    total_rows = 0
    total_saved = 0

    for fund_id, fund_data in funds.items():
        fund_name = fund_data["name"]
        rows = fund_data["rows"]

        # Check latest date already in database for this fund
        latest_date = database.get_latest_date(conn, fund_id)

        if latest_date:
            # Filter to only rows newer than latest saved date
            new_rows = [r for r in rows if r["date"] > latest_date]
            print(
                f"  {fund_name:<55} | latest in DB: {latest_date} | {len(new_rows)} new rows to add"
            )
        else:
            # No existing data — insert everything
            new_rows = rows
            dates = [r["date"] for r in rows]
            print(
                f"  {fund_name:<55} | first import | {len(new_rows)} rows | {min(dates)} → {max(dates)}"
            )

        if new_rows:
            saved = database.save_prices(
                conn, fund_id, fund_name, new_rows, asset_type="Fund"
            )
            total_saved += saved

        total_rows += len(new_rows)

    conn.close()

    print(
        f"\nDone. {total_saved} of {total_rows} new rows imported. {skipped} rows skipped."
    )


if __name__ == "__main__":
    main()
