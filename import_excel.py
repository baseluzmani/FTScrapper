# import_excel.py
# Imports historical fund data from the SL funds Excel sheets into the database.
# Can be run repeatedly — only new rows (after the latest saved date) are inserted.
#
# Expected sheet structure:
#   Column A: Date (DD/MM/YYYY text or Excel date)
#   Column B: Open
#   Column C: High
#   Column D: Low
#   Column E: Close
#   Column F: Volume
#   Column J: FT URL (row 1) / Fund Name (row 2) / Fund ID (row 3)
#
# Usage:
#   python3 import_excel.py                          # reads Funds_Database.xlsx
#   python3 import_excel.py ~/Downloads/myfile.xlsx  # specific file

import sys
import openpyxl
from datetime import datetime
import database

DEFAULT_FILE = "Funds_Database.xlsx"

# Map each Excel sheet name to its fund ID in the database
SHEET_TO_FUND = {
    "SL Asia": "GB0031728438:GBP",
    "SL Gold": "GB00B3VNFD68:GBP",
    "SL China": "GB00B7Z71453:GBP",
    "SL JPM NatRes": "GB00B61JR401:GBP",
    "SL Ishares Pacific": "GB00B849FB47:GBP",
    "SL Infrastructure": "GB00BF0TZK67:GBP",
    "SL Macquarie Global Infr": "GB00B3K5WJ87:GBP",
}


def parse_date(value):
    """Parse a date from either a datetime object or a DD/MM/YYYY string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def get_fund_name(ws, fund_id):
    """
    Try to extract the fund name from column J rows 1-3.
    Falls back to fund_id if not found.
    """
    for row_idx in range(1, 4):
        for col_idx in range(9, 12):  # columns J, K, L
            cell = ws.cell(row=row_idx, column=col_idx).value
            if (
                cell
                and isinstance(cell, str)
                and len(cell) > 5
                and "http" not in cell
                and ":" not in cell
            ):
                return cell.strip()
    return fund_id


def import_sheet(ws, fund_id, conn):
    """Read all data rows from a sheet and save new ones to the database."""
    fund_name = get_fund_name(ws, fund_id)

    rows = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if row[0] is None:
            skipped += 1
            continue

        date_formatted = parse_date(row[0])
        if not date_formatted:
            skipped += 1
            continue

        # Skip header row if it somehow appears in data
        if str(row[0]).strip().lower() == "date":
            continue

        try:
            rows.append(
                {
                    "date": date_formatted,
                    "open": float(row[1] or 0),
                    "high": float(row[2] or 0),
                    "low": float(row[3] or 0),
                    "close": float(row[4] or 0),
                    "volume": int(float(row[5] or 0)),
                }
            )
        except (TypeError, ValueError):
            skipped += 1
            continue

    # Only save rows newer than what's already in the database
    latest_date = database.get_latest_date(conn, fund_id)
    if latest_date:
        new_rows = [r for r in rows if r["date"] > latest_date]
    else:
        new_rows = rows

    saved = database.save_prices(conn, fund_id, fund_name, new_rows, asset_type="Fund")
    return len(rows), saved, skipped


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    print(f"Opening: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    conn = database.get_connection()
    database.create_table(conn)

    total_rows = 0
    total_saved = 0

    for sheet_name, fund_id in SHEET_TO_FUND.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        rows, saved, skipped = import_sheet(ws, fund_id, conn)
        total_rows += rows
        total_saved += saved

        print(
            f"  {sheet_name:<30} | {rows:4} rows found | {saved:4} saved | {skipped} skipped"
        )

    conn.close()
    print(f"\nDone. {total_saved} of {total_rows} rows imported.")


if __name__ == "__main__":
    main()
